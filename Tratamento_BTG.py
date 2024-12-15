### 0. Importação de bibliotecas
from pathlib import Path
import requests
import pandas as pd
from openpyxl import load_workbook
import re
import numpy as np

### 1. Definição de funções
def rename_column(col):
    """Remove sufixos desnecessários (.0) e formata colunas."""
    col = str(col).strip()
    col = re.sub(r'\.0$', '', col)  # Remove sufixo .0 no final
    match = re.match(r'(\dQ)\s?(\d{4})', col)
    if match:
        return f"{match.group(1)}{match.group(2)[-2:]}"  # Ajusta nomes como "4Q 2011"
    return col

# Função para formatar apenas valores numéricos
def transforma_float_str(x):
    try:
        return f"{float(x):.4f}".replace('.', ',')
    except ValueError:
        return x  # Retorna o valor original se não for numérico

def formata_tabelas(df, sheet_name):
    """Função para formatar a planilha bruta importada via API."""

    # Ajustar o header do dataframe
    df.columns = df.iloc[0]  # Atribuir a primeira linha como cabeçalho
    df = df[1:]  # Remover a linha do cabeçalho original
    df.dropna(axis=1, how='all', inplace=True)

    # Renomear colunas para remover problemas como sufixo .0 e formata as com formato #Q####
    df.columns = [rename_column(col) for col in df.columns]

    # Criar coluna para categoria superior
    mascara_categoria_superior = df.iloc[:, 1:].isna().all(axis=1)
    df['Categoria Superior'] = df.loc[mascara_categoria_superior, df.columns[0]]
    df['Categoria Superior'] = df['Categoria Superior'].fillna(method='ffill')
    df = df[~df['Categoria Superior'].isna()]
    df = df.loc[~mascara_categoria_superior]

    # Criar coluna de referencia
    df['Referencia'] = df['Categoria Superior'] + ' - ' + df[df.columns[0]]
    df.drop(columns=[df.columns[0], 'Categoria Superior'], inplace=True)

    # Adicionar coluna de referência da sheet
    ref = df.pop('Referencia')
    df.insert(0, 'Referencia', ref)
    df.insert(1, 'Sheet', sheet_name)

    # Identificar colunas numéricas para transformação
    colunas_identificadores = ['Referencia', 'Sheet'] #Sempre as mesmas
    colunas_valores = [col for col in df.columns if col not in colunas_identificadores]
    
    #Não tava dando certo aqui mas acho que pode botar no base. Teoricamente os valores são sempre numeros e se não forem nao devem incluir pontuação
    #Substitui separador "." para vírgula para facilitar intepretação no Excel; também formata para 2 casas decimais para facilitar visualização
    #colunas_numericas = df.select_dtypes(include='number').columns
    #df[colunas_numericas] = df[colunas_numericas].applymap(lambda x: f"{x:.2f}".replace(".", ","))


    # Transformar em formato tidy
    tidy_df = pd.melt(
        df,
        id_vars=colunas_identificadores,
        value_vars=colunas_valores,
        var_name='Tipo',
        value_name='Valor'
    )

    # Substituir valores NaN por '-' para visualização
    tidy_df['Valor'] = tidy_df['Valor'].replace(np.nan, '-')

    return tidy_df

### 2. Parâmetros gerais
url = 'https://api.mziq.com/mzfilemanager/v2/d/0afe1b62-e299-4dec-a938-763ebc4e2c11/79c66cca-4d48-0e24-db90-67450e78b597?origin=1'
cod_sucesso = 200
diretorio_export = str(Path.cwd())
nome_arquivo = 'Base_BTG'

### 3. Importação e tratamento
response = requests.get(url)
if response.status_code == cod_sucesso:
    with open("temp.xlsx", "wb") as f:
        f.write(response.content)
else:
    print(f"Erro ao baixar o arquivo: {response.status_code}")
    exit()

# Carregar nomes das sheets
workbook = load_workbook("temp.xlsx")
sheets = workbook.sheetnames
del workbook

dfs = {}
base_tidy = pd.DataFrame()
for sheet_name in sheets:
    dfs[sheet_name] = pd.read_excel("temp.xlsx", sheet_name=sheet_name)
    base_tidy = pd.concat([base_tidy, formata_tabelas(dfs[sheet_name], sheet_name)])
    
# Aplicar correçao de valor (muda . para , para ser interpretado no excel)
#Sempre na coluna "Valor" da base, que não muda, então sempre a ultima coluna é "Valor"
base_tidy['Valor'] = base_tidy['Valor'].apply(transforma_float_str)

### 4. Exportação

#Exportar a base tidy
base_tidy.to_csv(
    rf"{diretorio_export}\{'Base_BTG_Tidy'}.csv",
    index=False,
    encoding="utf-8-sig",
    sep=';',
    decimal=','
)

### Deletar o arquivo temporário
temp_file = Path("temp.xlsx")
if temp_file.exists():
    temp_file.unlink()
    print("Arquivo temporário 'temp.xlsx' deletado com sucesso.")
else:
    print("Arquivo temporário 'temp.xlsx' não encontrado para deletar.")
